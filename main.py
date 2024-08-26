from colorama import Fore
from pystyle import Colors, Box, Write, Center, Colorate
from os import name, system
from pystyle import Colorate
import os
import sys
import csv
import requests
import openpyxl
import random
import time
from faker import Faker

fake = Faker()


def console_clear():
    if os.sys.platform == 'win32':
        os.system('cls')
    else:
        os.system("clear")

def Continue():
    input('Нажмите Enter чтобы продолжить ')
    console_clear()
    software()

def search_db(search_s, d_path):
    for root, dirs, files in os.walk(d_path):
        for file in files:
            file_path = os.path.join(root, file)
            if file.endswith('.csv'):
                with open(file_path, 'r', encoding='utf-8') as csvfile:
                    reader = csv.reader(csvfile)
                    for row in reader:
                        if search_s in row:
                            print(f"Найдено в файле {file_path}, строка: {row}")
            elif file.endswith('.xlsx'):
                wb = openpyxl.load_workbook(file_path)
                for sheet in wb.sheetnames:
                    ws = wb[sheet]
                    for row in ws.iter_rows(values_only=True):
                        if search_s in row:
                            print(f"Найдено в файле {file_path}, лист {sheet}, строка: {row}")
            elif file.endswith('.txt'):
                with open(file_path, 'r', encoding='utf-8') as txtfile:
                    for line in txtfile:
                        if search_s in line:
                            print(f"Найдено в файле {file_path}, строка: {line.strip()}")

def ip_lookup(ip_address): 
  url = f"http://ip-api.com/json/{ip_address}"
  try:
    response = requests.get(url)
    data = response.json()
    if data.get("status") == "fail":
      return f"Ошибка: {data['message']}\n"
    
    info = ""
    for key, value in data.items():
      info += f"  |{key}: {value}\n"
    return info
  
  except Exception as e:
    return f"Произошла ошибка: {str(e)}\n"
  

def generate_Bday():
    data_1 = random.randint(1, 29)
    data_2 = random.randint(1, 12)
    data_3 = random.randint(1980, 2008)

    Bday = f'{data_1}.{data_2}.{data_3}'
    print(f"Ваша дата рождения: {Bday}")
    Continue()


def generate_passport():
        grad_start = ('Генерация началсь, подождите...')
        print(grad_start)
        time.sleep(3)

        fio_r = random.randint(1, 10)
        if fio_r == 1:
            fio = 'Мишенко Виктор Михайлович'
        elif fio_r == 2:
            fio = 'Сухонов Михаил Артемович'
        elif fio_r == 3:
            fio = 'Геодозов Виталий Юрьевич'
        elif fio_r == 4:
            fio = 'Шабалин Кирилл Витальевич'
        elif fio_r == 5:
            fio = 'Кострюков Евгений Петрович'
        else: 
            fio = 'Шевченко Матвей Евгеньевич'

        bd_data_1 = random.randint(1, 29) # День
        if bd_data_1 <= 9:
            bd_data_1 = f'0{bd_data_1}'
        bd_data_2 = random.randint(1, 12) # Месяц
        if bd_data_2 <= 9:
            bd_data_2 = f'0{bd_data_2}'
        bd_data_3 = random.randint(1980, 2009) # Год рождения
        Bday = f'{bd_data_1}.{bd_data_2}.{bd_data_3}'


        City_Bday_r = random.randint(1, 10)
        if City_Bday_r == 1:
            City_Bday = 'Москва'
        elif City_Bday_r == 2:
            City_Bday = 'Новокузнецк'
        elif City_Bday_r == 3:
            City_Bday = 'Пенза'
        elif City_Bday_r == 4:
            City_Bday = 'Санкт-Петербург'
        elif City_Bday_r == 5:
            City_Bday = 'Калининград'
        else: 
            City_Bday = 'Кривой рог'


        cr_data_1 = random.randint(100, 999)
        cr_data_2 = random.randint(100, 999)
        Code_razdel = f'{cr_data_1}-{cr_data_2}'


        if City_Bday_r == 1:
            Who_give = 'ГУ МВД ПО МОСКОВСКОЙ ОБЛАСТИ'
        elif City_Bday_r == 2:
            Who_give = 'ГУ МВД ПО НОВОКУЗНЕЦКУ ОБЛАСТИ' 
        elif City_Bday_r == 3:
            Who_give = 'ГУ МВД ПО ПЕНЗОВСКОЙ ОБЛАСТИ'
        elif City_Bday_r == 4:
            Who_give = 'ГУ МВД ПО САНКТ-ПЕТЕРБУРГА ОБЛАСТИ'
        elif City_Bday_r == 5:
            Who_give = 'ГУ МВД ПО КАЛИНИНГРАДА ОБЛАСТИ'



        # if bd_data_2 >= 11:
        #     When_give_2 = bd_data_2 + 1
        # if bd_data_2 >= 12:
        #     When_give_2 = bd_data_2
        # else:
        #     When_give_2 = bd_data_2 + 1
        When_give_1 = bd_data_1
        When_give_2 = bd_data_2
        When_give_3 = bd_data_3 + 14

        When_give = f'{When_give_1}.{When_give_2}.{When_give_3}'

        M_or_F_r = random.randint(1, 2)
        if M_or_F_r == 1:
            M_or_F = 'Мужской'
        elif M_or_F_r == 2:
            M_or_F = 'Женский'

        series_Passport_1 = random.randint(45, 58)
        series_Passport_2 = random.randint(4, 9)
        if series_Passport_2 <= 9:
            series_Passport_2 = f'0{series_Passport_2}'
        series_Passport = f'{series_Passport_1} {series_Passport_2}'
        
        Number_Passport = random.randint(100000, 900000)
        
        passport = f'''
╭──────────────────────[FAKE]───────────────────────╮
 ФИО              : {fio}                                        
 Дата рождения    : {Bday}                    
 Город рождения   : {City_Bday}
 Код подразделения: {Code_razdel}
 Кем выдан        : Сам придумай
 Когда выдан      : {When_give}
 Пол              : {M_or_F}
 Серия паспорта   : {series_Passport}
 Номер паспорта   : {Number_Passport}                           
╰──────────────────────[FAKE]───────────────────────╯

'''
        print(passport)
        Continue()

def print_number_info(self):
        user_number = input("Введите номер телефона (например, +79833170773): ").strip()

        if user_number:
            print("Поиск данных...\n")
            data = self.get_number_data(user_number)

            if data.get("status_error"):
                print("Ошибка: Не удалось получить данные. Проверьте номер телефона и попробуйте снова.")
                return

            if data.get("limit") == 0:
                print("Вы израсходовали все лимиты запросов.")
                return

            country = data.get('country', {})
            region = data.get('region', {})
            other = data.get('0', {})

            print(f"Страна: {country.get('name', 'Не найдено')}, {country.get('fullname', 'Не найдено')}")
            print(f"Город: {other.get('name', 'Не найдено')}")
            print(f"Почтовый индекс: {other.get('post', 'Не найдено')}")
            print(f"Оператор: {other.get('oper', 'Не найдено')}, {other.get('oper_brand', 'Не найдено')}, {other.get('def', 'Не найдено')}")
            print(f"Местоположение: {country.get('name', 'Не найдено')}, {region.get('name', 'Не найдено')}, {other.get('name', 'Не найдено')} ({region.get('okrug', 'Не найдено')})")

            latitude = other.get('latitude', 'Не найдено')
            longitude = other.get('longitude', 'Не найдено')
            location = data.get('location', 'Не найдено')
            lang = country.get('lang', 'Не найдено').title()
            lang_code = country.get('langcod', 'Не найдено')
            capital = data.get('capital', {}).get('name', 'Не найдено')

            print(f"Открыть на карте (google): https://www.google.com/maps/place/{latitude}+{longitude}")
            print(f"Локация: {location}")
            print(f"Край/Округ/Область: {region.get('name', 'Не найдено')}, {region.get('okrug', 'Не найдено')}")
            print(f"Столица: {capital}")
            print(f"Широта/Долгота: {latitude}, {longitude}")
        else:
            print("Ошибка: Номер телефона не введен.")

def print_bots():
    bots = '''
1. @phonenumberinformation_bot
2. @Quick_osintik_bot
3. @UniversalSearchRobot
4. @search_himera_bot
5. @Solaris_Search_Bot
6. @Zernerda_bot
7. @t_sys_bot
8. @OSINTInfoRobot
9. @LBSE_bot
10. @SovaAppBot
11. @poiskorcombot
12. @SEARCHUA_bot
13. @helper_inform_bot
14. @infobazaa_bot
15. @declassified_bot
16. @GHack_search_bot
17. @osint_databot
18. @Informator_BelBot
19. @HowToFindRU_Robot
20. @SEARCH2UA_bot
21. @UsersSearchBot
22. @BITCOlN_BOT
23. @ce_poshuk_bot
24. @BlackatSearchBot
25. @dataisbot
26. @n3fm4xw2rwbot
27. @cybersecdata_bot
28. @safe_search_bot
29. @getcontact_real_bot
30. @PhoneLeaks_bot
31. @useridinfobot 
32. @mailcat_s_bot
33. @last4mailbot
34. @holehe_s_bot
35. @bmi_np_bot
36. @clerkinfobot
37. @kolibri_osint_bot
38. @getcontact_premium_bot
39. @phone_avito_bot
40. @pyth1a_0racle_bot
41. @olx_phone_bot
42. @ap_pars_bot
43. @SPOwnerBot
44. @regdatebot
45. @ibhldr_bot
46. @TgAnalyst_bot
47. @cryptoscanning_bot
48. @LinkCreatorBot
49. @telesint_bot
50. @Checknumb_bot
51. @TelpoiskBot_bot
52. @TgDeanonymizer_bot
53. @protestchat_bot
54. @locatortlrm_bot
55. @GetCont_bot
56. @usinfobot
57. @SangMataInfo_bot
58. @creationdatebot
59. @tgscanrobot
60. @InfoVkUser_bot
61. @getfb_bot
62. @GetOK2bot
63. @VKHistoryRobot
64. @detectiva_robot
65. @FindNameVk_bot
66. @vk2017robot
67. @AgentFNS_bot
68. @OpenDataUABot
69. @egrul_bot
70. @Bumz639bot
80. @ogrn_bot
90. @ShtrafKZBot
91. @egrnrobot
92. @VipiskaEGRNbot
93. @Search_firm_bot
94. @geomacbot
95. @pwIPbot
96. @ipscorebot
97. @ip_score_checker_bot
98. @FakeSMI_bot
99. @ipinfo_check_bot
100. @Search_IPbot
101. @WhoisDomBot
102. @vimebasebot
103. @maigret_osint_bot
104. @PasswordSearchBot
105. @ddg_stresser_bot
106. @BotAvinfo_bot
107. @reverseSearch2Bot
108. @pimeyesbot
109. @findfacerobot
110. @CarPlatesUkraineBot
111. @nomerogrambot
112. @ShtrafyPDRbot
113. @cerbersearch_bot
'''

    print(bots)
    Continue()

def generate_card():
        grad_start = ('Генерация началсь, подождите...')
        print(grad_start)
        time.sleep(3)

        card_data_1_r = random.randint(1, 2)
        if card_data_1_r == 1:
            card_data_1 = '2202'
        elif card_data_1_r == 2:
            card_data_1 = '2200'
        
        card_data_2 = random.randint(1000, 9999)
        card_data_3 = random.randint(1000, 9999)
        card_data_4 = random.randint(1000, 9999)

        card_number = f"{card_data_1} {card_data_2} {card_data_3} {card_data_4}"



        fio_r = random.randint(1, 10)
        fio_r = random.randint(1, 10)
        if fio_r == 1:
            fio = 'Мишенко Виктор Михайлович'
        elif fio_r == 2:
            fio = 'Сухонов Михаил Артемович'
        elif fio_r == 3:
            fio = 'Геодозов Виталий Юрьевич'
        elif fio_r == 4:
            fio = 'Шабалин Кирилл Витальевич'
        elif fio_r == 5:
            fio = 'Кострюков Евгений Петрович'
        else: 
            fio = 'Шевченко Матвей Евгеньевич'

        srok_1 = random.randint(1, 12)
        srok_2 = random.randint(25, 34)
        if srok_1 <= 9:
            srok_1 = f'0{srok_1}'
        srok = f"{srok_1}/{srok_2}" 

        cvc = random.randint(100, 999)

        bank_r = random.randint(1, 4)
        if bank_r == 1:
            bank = 'Т-Банк'
        elif bank_r == 2:
            bank = 'Сбербанк'
        elif bank_r == 3:
            bank = 'ВТБ'
        elif bank_r == 4:
            bank = 'Совкомбанк'


        card = f'''
╭──────────────────────[FAKE]───────────────────────╮
 Владелер карты     : {fio}
 Номер карты        : {card_number}
 Срок действия карты: {srok}
 CVC код карты      : {cvc}
 Банк               : {bank}                          
╰──────────────────────[FAKE]───────────────────────╯

'''
        print(card)
        Continue()    
    
def generate_name():
    select_who = input('М или Ж -> ')
    if select_who == 'М':
        grad_print = (fake.name_male())
        print(f'Вот ваше имя: {grad_print}')
        Continue()
    elif select_who == 'Ж':
        grad_print = (fake.name_female())
        print(f'вот ваше имя: {grad_print}')
        Continue()

def generate_address():
        City_Bday_r = random.randint(1, 10)
        if City_Bday_r == 1:
            City = 'Москва'
        elif City_Bday_r == 2:
            City = 'Новокузнецк'
        elif City_Bday_r == 3:
            City = 'Пенза'
        elif City_Bday_r == 4:
            City = 'Санкт-Петербург'
        elif City_Bday_r == 5:
            City = 'Калининград'
        else: 
            City = 'Кривой рог'
            
        Street_r = random.randint(1, 3)
        if Street_r == 1:
            Street = 'Ленинская'
        elif Street_r == 2:
            Street = 'Михайловская'
        elif Street_r == 3:
            Street = 'Петрозаводская'

        N_house = random.randint(1, 100)
    
        Appartament = random.randint(1, 400)





        address = f'''
╭──────────────────────[FAKE]───────────────────────────────────────╮
 Город: {City}
 Улица: {Street}   
 Дом: {N_house}
 Квартира: {Appartament}
-----------------------
 Адрес строкой: {City}, улица {Street}, дом {N_house}, кв. {Appartament}
╰──────────────────────[FAKE]───────────────────────────────────────╯

'''
        print(address)
        Continue()


banner = '''
▀███▀▀▀██▄            ▀███        ▄▄█▀▀██▄  
  ██    ▀██▄            ██      ▄██▀    ▀██▄
  ██     ▀█████  ▀███   ██  ▄██▀██▀      ▀██
  ██      ██ ██    ██   ██ ▄█   ██        ██
  ██     ▄██ ██    ██   ██▄██   ██▄      ▄██
  ██    ▄██▀ ██    ██   ██ ▀██▄ ▀██▄    ▄██▀
▄████████▀   ▀████▀███▄████▄ ██▄▄ ▀▀████▀▀  

                dev: @v1nx_dev
'''

menu = '''
╔                                             ╗
  1. Поиск по базе      | 6. Боты
  2. Поиск по IP        | 7. Генератор карты
  3. Поиск по номеру    | 8. Генератор имени
  4. Генератор паспорта | 9. Генератор адреса
  5. Генератор др       | 10. выход
╚                                             ╝
'''
def software():
    console_clear()
    grad_banner = Colorate.Horizontal(Colors.blue_to_cyan, Center.XCenter(banner))
    grad_menu = Colorate.Horizontal(Colors.blue_to_cyan, Center.XCenter(menu))

    print(grad_banner)
    print(grad_menu)

    choice = input('Выберите опцию -> ')

        

    if choice == '1':
        search_s = input("Введите строку для поиска: ")
        d_path = input("Введите путь к папке для поиска: ")
        search_db(search_s, d_path)

    if choice == '2':
        ip_address = input('Введите IP -> ')
        result = ip_lookup(ip_address)
        print(f'\n{result}')
        Continue()

    if choice == '3':
        search_s = input("Введите строку для поиска: ")
        d_path = input("Введите путь к папке для поиска: ")
        search_db(search_s, d_path)
    
    if choice == '4':
        generate_passport()
    
    if choice == '5':
        generate_Bday()
    
    if choice == '6':
        print_bots()
    
    if choice == '7':
        generate_card()

    if choice == '8':
        generate_name()
    
    if choice == '9':
        generate_address()

    if choice == '10':
        sys.exit()

    else:
        print('Неправильный ввод')

software()
